import openpyxl
import numpy as np
from Bio import Phylo
from UPGMA import UPGMA
from CMM import CMM
from matplotlib import pyplot as plt

# CONSTANTS
noisily = True
file_path = "Data/Hw1-words.xlsx"
UPGMA_plot_path = "Plots/UPGMA Tree.pdf"
CMM_plot_path = "Plots/CMM Tree.pdf"
data_cache_path = "Data/Cache"


# HELPER FUNCTIONS
def tree_to_newick(t) -> str:
    if t.right is None:
        return t.left + ":" + str(t.uh)
    else:
        return (
                "("
                + ",".join([tree_to_newick(x) for x in [t.left, t.right]])
                + "):"
                + str(t.uh)
        )


def read_excel_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active # !!! CONSIDER REVISING THIS LINE !!!
    languages = []
    words = []
    for row in range(1, sheet.max_row + 1):  # Iterate over rows starting from the first row
        language = sheet.cell(row=row, column=1).value
        if language:  # Check if language name is not None
            language = language.replace("(", "")
            language = language.replace(")", "")
            language = language.replace(" ", "-")
            languages.append(language)
            word_list = [sheet.cell(row=row, column=col).value for col in range(2, sheet.max_column + 1)]
            words.append(word_list)
    return languages, words


def GenerateUpgma(languages, words):
    if noisily: print("!!! START GENERATING UPGMA TREE !!!")
    if noisily: print("Generating hamming distances matrix...")
    distances = np.zeros((len(words), len(words)), dtype=int)
    for i in range(len(words)):
        for j in range(i + 1, len(words)):
            dist = sum(1 for w1, w2 in zip(words[i], words[j]) if w1 != w2)
            distances[i, j] = dist
            distances[j, i] = dist
    if noisily: print("Building UPGMA tree...")
    upgma_tree = UPGMA(distances, languages).tree
    if noisily: print("Converting UPGMA tree to newick format...")
    upgma_newick = tree_to_newick(upgma_tree)
    if noisily: print(f"Newick formatted tree: {upgma_newick}")
    if noisily: print("Writing data to cache...")
    with open(data_cache_path, 'w') as file:
        file.write(upgma_newick)
    if noisily: print("Converting cache data to Phylo tree...")
    upgma_p_tree = Phylo.read(data_cache_path, "newick")
    if noisily: print(f"Plylo tree: {upgma_p_tree}")
    if noisily: print("Drawing UPGMA tree...")
    Phylo.draw(upgma_p_tree, do_show=False)
    xmin, xmax, ymin, ymax = plt.axis()
    # 0.5 is a magic number used to increase bondary to fit longer language names.
    plt.axis([xmin, xmax + 0.5, ymin, ymax])
    if noisily: print(f"Saving UPGM tree to file system at {UPGMA_plot_path}.")
    plt.savefig(UPGMA_plot_path)
    if noisily: print("!!! FINISHED GENERATING UPGMA TREE !!!")


def GenerateCmm(languages, words):
    if noisily: print("!!! START GENERATING CMM TREE !!!")
    if noisily: print("Initialize data structures...")
    comm_mutation = []
    transposed_matrix = list(map(list, zip(*words)))
    if noisily: print("Transpose the input matrix...")
    for col in transposed_matrix:
        word_count = {}
        # Count the occurrences of each word in the column
        for word in col:
            word_count[word] = word_count.get(word, 0) + 1
        # Find the most frequent word in the column
        most_frequent_word = max(word_count, key=word_count.get)
        comm_mutation.append(most_frequent_word)
    if noisily: print("Generate mutation set...")
    set_of_mutations = [[] for i in range(len(languages))]
    for i in range(len(words)):
        for j in range(len(words[i])):
            if (words[i][j] != comm_mutation[j]):
                set_of_mutations[i].append(str(j + 1) + words[i][j])
    if noisily: print("Build the CMM matrix...")
    cmm = [[None for i in range(len(languages))] for j in range(len(languages))]
    for i in range(len(set_of_mutations)):
        for j in range(i, len(set_of_mutations)):
            common_list = list(set(set_of_mutations[i]) & set(set_of_mutations[j]))
            cmm[i][j] = common_list
    if noisily: print("Build the CMM tree...")
    cmm_tree = CMM(cmm, languages).tree
    if noisily: print("Convert the CMM to newick format...")
    # Given that the topmost node in the common mutation list has potentially n children,
    # while the remaining nodes have at most 2 children, the following code iterates through
    # each child of the topmost node (in reverse order) and prints the Newick representation of its subtree.
    cmm_newick = "(" + ",".join([tree_to_newick(tree) for tree in cmm_tree]) + ")"
    if noisily: print(f"Newick formatted tree: {cmm_newick}")
    if noisily: print("Writing data to cache...")
    with open(data_cache_path, 'w') as file:
        file.write(cmm_newick)
    if noisily: print("Converting cache data to Phylo tree...")
    cmm_p_tree = Phylo.read(data_cache_path, "newick")
    if noisily: print(f"Plylo tree: {cmm_p_tree}")
    if noisily: print("Drawing CMM tree...")
    Phylo.draw(cmm_p_tree, do_show=False)
    xmin, xmax, ymin, ymax = plt.axis()
    # 0.5 is a magic number used to increase bondary to fit longer language names.
    plt.axis([xmin, xmax + 0.5, ymin, ymax])
    if noisily: print(f"Saving CMM tree to file system at {CMM_plot_path}.")
    plt.savefig(CMM_plot_path)
    if noisily: print("!!! FINISHED GENERATING CMM TREE !!!")


# ENTRY/EXIT POINT
if __name__ == "__main__":
    if noisily: print()
    if noisily: print("****************************************************")
    if noisily: print("****************************************************")
    if noisily: print("*** STARING COMPUTATIONAL LINGUISTICS HOMEWORK 1 ***")
    if noisily: print("****************************************************")
    if noisily: print("****************************************************")
    if noisily: print()
    if noisily: print(f"Reading excel input file at {file_path}")
    languages, words = read_excel_file(file_path)
    GenerateUpgma(languages, words)
    if noisily: print()
    GenerateCmm(languages, words)
    if noisily: print()
    if noisily: print("****************************************************")
    if noisily: print("****************************************************")
    if noisily: print("***   FINISHED COMPUTATIONAL LINGUISTICS HW 1    ***")
    if noisily: print("****************************************************")
    if noisily: print("****************************************************")

